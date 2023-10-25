# excelapp/views.py
from django.shortcuts import render
from django.http import HttpResponse
from openpyxl import Workbook
from .forms import ExcelFormSet, TableConfigForm
from .sum_forms import ExcelSumFormSet
from openpyxl.utils import get_column_letter
from io import BytesIO
from reportlab.lib.pagesizes import letter, landscape, portrait, A4
from reportlab.pdfgen import canvas
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, PageBreak
from reportlab.lib import colors
from django.http import JsonResponse
from openpyxl.styles import Font
# from openpyxl.worksheet.merge import merge_cells
# from openpyxl import Workbook


def generate_excel(request):
    if request.method == 'POST':
        form = TableConfigForm(request.POST)
        if form.is_valid():
            num_rows = form.cleaned_data['num_rows']
            num_cols = form.cleaned_data['num_cols']
            cell_width = form.cleaned_data['cell_width']
            cell_height = form.cleaned_data['cell_height']

            # Create a new Workbook and select the active sheet
            wb = Workbook()
            ws = wb.active

            # Set the font size and bold for the entire sheet
            bold_font = Font(bold=True)
            for row in ws.iter_rows():
                for cell in row:
                    cell.font = bold_font

            # Create an empty table with the specified number of rows and columns
            for row in range(1, num_rows + 1):
                for col in range(1, num_cols + 1):
                    cell = ws.cell(row=row, column=col)
                    cell.value = f'Row {row}, Col {col}'

            # Adjust column widths and row heights
            for col_num in range(1, num_cols + 1):
                col_letter = get_column_letter(col_num)
                col_width = cell_width
                ws.column_dimensions[col_letter].width = col_width

            for row_num in range(1, num_rows + 1):
                row_height = cell_height
                ws.row_dimensions[row_num].height = row_height

            # Create a response with the Excel file
            response = HttpResponse(content_type='application/ms-excel')
            response['Content-Disposition'] = 'attachment; filename="empty_table.xlsx"'

            # Save the workbook to the response
            wb.save(response)

            return response
    else:
        form = TableConfigForm()

    return render(request, 'excelapp/table_config.html', {'form': form})




def generate_pdf(request):
    if request.method == 'POST':
        form = TableConfigForm(request.POST)
        if form.is_valid():
            num_rows = form.cleaned_data['num_rows']
            num_cols = form.cleaned_data['num_cols']
            cell_width = form.cleaned_data['cell_width']
            cell_height = form.cleaned_data['cell_height']

            # Create a PDF document
            pdf_buffer = BytesIO()
            doc = SimpleDocTemplate(pdf_buffer, pagesize=(portrait(A4)))  #landscape(letter)

            # Create an empty table with the specified number of rows and columns
            data = [['' for _ in range(num_cols)] for _ in range(num_rows)]
            table = Table(data, colWidths=[cell_width] * num_cols, rowHeights=[cell_height] * num_rows)

            # Set custom style for the table
            table.setStyle(TableStyle([
                ('INNERGRID', (0, 0), (-1, -1), 0.5, (0, 0, 0)),  # Add inner grid lines
                ('BOX', (0, 0), (-1, -1), 0.5, (0, 0, 0))  # Add cell borders
            ]))

            # Build the PDF
            elements = [table]
            doc.build(elements)

            # Serve the PDF as a response
            pdf_buffer.seek(0)
            response = HttpResponse(pdf_buffer.read(), content_type='application/pdf')
            response['Content-Disposition'] = 'attachment; filename="empty_table.pdf"'
            return response
    else:
        form = TableConfigForm()

    return render(request, 'excelapp/table_config.html', {'form': form})





def create_pdf(request):
    if request.method == 'POST':
        formset = ExcelFormSet(request.POST)

        if formset.is_valid():
            # Create a list to hold the data
            data = [['Name', 'Age', 'City']]

            for form in formset:
                data.append([form.cleaned_data['name'], form.cleaned_data['age'], form.cleaned_data['city']])

            # Create a PDF buffer and a PDF document
            buffer = BytesIO()
            doc = SimpleDocTemplate(buffer, pagesize=portrait(A4))

            # Create a table from the data and set style
            table = Table(data)
            table.setStyle(TableStyle([
                ('BACKGROUND', (0, 0), (-1, 0), colors.grey),
                ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
                
                ('ALIGN', (0, 0), (0, 0), 'CENTER'),
                ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                ('BOTTOMPADDING', (0, 0), (-1, 0), 5),
                
                ('BACKGROUND', (0, 1), (-1, -1), colors.white),
                ('GRID', (0, 0), (-1, -1), 1, colors.black),
            ]))

            # Add the table to the PDF document
            elements = []
            elements.append(table)

            doc.build(elements)

            # Close the PDF response and return
            buffer.seek(0)
            response = HttpResponse(buffer.read(), content_type='application/pdf')
            response['Content-Disposition'] = 'attachment; filename="example.pdf"'
            return response

    else:
        formset = ExcelFormSet()

    return render(request, 'excelapp/create_pdf.html', {'formset': formset})


def create_sum_pdf(request):
    if request.method == 'POST':
        formset = ExcelSumFormSet(request.POST)

        if formset.is_valid():
            # Create a list to hold the data
            data = [['Question', 'Answer']]

            for form in formset:
                data.append([form.cleaned_data['question'], form.cleaned_data['answer']])

            # Create a PDF buffer and a PDF document
            buffer = BytesIO()
            doc = SimpleDocTemplate(buffer, pagesize=portrait(A4))

            # Create a table from the data and set style
            table = Table(data)
            table.setStyle(TableStyle([
                ('BACKGROUND', (0, 0), (-1, 0), colors.grey),
                ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
                
                ('ALIGN', (0, 0), (0, 0), 'CENTER'),
                ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                ('BOTTOMPADDING', (0, 0), (-1, 0), 5),
                
                ('BACKGROUND', (0, 1), (-1, -1), colors.white),
                ('GRID', (0, 0), (-1, -1), 1, colors.black),
            ]))

            # Add the table to the PDF document
            elements = []
            elements.append(table)

            doc.build(elements)

            # Close the PDF response and return
            buffer.seek(0)
            response = HttpResponse(buffer.read(), content_type='application/pdf')
            response['Content-Disposition'] = 'attachment; filename="sum_example.pdf"'
            return response

    else:
        formset = ExcelSumFormSet()

    return render(request, 'excelapp/create_sum_pdf.html', {'formset': formset})


















def create_excel(request):
    if request.method == 'POST':
        formset = ExcelFormSet(request.POST)

        if formset.is_valid():
            # Create an Excel workbook
            workbook = Workbook()
            sheet = workbook.active

            # Write headers
            headers = ['Name', 'Age', 'City']
            for col_num, header in enumerate(headers):
                sheet.cell(row=1, column=col_num + 1, value=header)

            # Write data to the sheet and adjust column widths
            for row_num, form in enumerate(formset, start=2):
                data = [form.cleaned_data['name'], form.cleaned_data['age'], form.cleaned_data['city']]
                for col_num, value in enumerate(data):
                    cell = sheet.cell(row=row_num, column=col_num + 1, value=value)
                    column_letter = get_column_letter(col_num + 1)
                    column_width = len(str(value)) + 2
                    if sheet.column_dimensions[column_letter].width is None or sheet.column_dimensions[column_letter].width < column_width:
                        sheet.column_dimensions[column_letter].width = column_width  # Adjust column width

            # Set row heights based on content
            for row in sheet.iter_rows(min_row=2, max_row=len(formset) + 1):
                max_height = max(
                    cell.alignment.vertical if cell.alignment.vertical is not None else 0
                    for cell in row
                )
                sheet.row_dimensions[row[0].row].height = max_height

            # Create an HTTP response for the Excel file
            response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
            response['Content-Disposition'] = 'attachment; filename="example.xlsx"'
            workbook.save(response)

            return response
    else:
        formset = ExcelFormSet()

    return render(request, 'excelapp/create_excel.html', {'formset': formset})
