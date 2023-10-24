# excelapp/views.py
from django.shortcuts import render
from django.http import HttpResponse
from openpyxl import Workbook
from .forms import ExcelFormSet
from .sum_forms import ExcelSumFormSet
from openpyxl.utils import get_column_letter
from io import BytesIO
from reportlab.lib.pagesizes import letter, landscape
from reportlab.pdfgen import canvas
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, PageBreak
from reportlab.lib.pagesizes import letter
from reportlab.lib import colors
# excelapp/views.py


def create_pdf(request):
    if request.method == 'POST':
        formset = ExcelFormSet(request.POST)

        if formset.is_valid():
            # Create a list to hold the data
            data = [['Name', 'Age', 'City']]
            
            for form in formset:
                if form.has_changed():  # Check if the form is changed (i.e., not empty)
                    data.append([form.cleaned_data['name'], form.cleaned_data['age'], form.cleaned_data['city']])

            # Create a PDF buffer and a PDF document
            buffer = BytesIO()
            doc = SimpleDocTemplate(buffer, pagesize=landscape(letter))

            # Create a table from the data and set style
            table = Table(data)
            table.setStyle(TableStyle([
                ('BACKGROUND', (0, 0), (-1, 0), colors.grey),
                ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
                ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
                ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
                ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
                ('GRID', (0, 0), (-1, -1), 1, colors.black),
            ]))

            # Add the table to the PDF document
            elements = []
            elements.append(table)

            doc.build(elements)

            # Close the PDF response and return
            buffer.seek(0)
            response = HttpResponse(buffer.read(), content_type='application/pdf')
            response['Content-Disposition'] = 'attachment; filename="example.pdf"'
            buffer.close()

            return response

    else:
        formset = ExcelFormSet()

    return render(request, 'excelapp/create_pdf.html', {'formset': formset})


def create_sum_pdf(request):
    if request.method == 'POST':
        formset = ExcelSumFormSet(request.POST)

        if formset.is_valid():
            # Create a list to hold the data
            data = [['Question', 'Answer']]

            for form in formset:
                data.append([form.cleaned_data['question'], form.cleaned_data['answer']])

            # Create a PDF buffer and a PDF document
            buffer = BytesIO()
            doc = SimpleDocTemplate(buffer, pagesize=landscape(letter))

            # Create a table from the data and set style
            table = Table(data)
            table.setStyle(TableStyle([
                ('BACKGROUND', (0, 0), (-1, 0), colors.grey),
                ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
                
                ('ALIGN', (0, 0), (0, 0), 'CENTER'),
                ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                ('BOTTOMPADDING', (0, 0), (-1, 0), 5),
                
                ('BACKGROUND', (0, 1), (-1, -1), colors.white),
                ('GRID', (0, 0), (-1, -1), 1, colors.black),
            ]))

            # Add the table to the PDF document
            elements = []
            elements.append(table)

            doc.build(elements)

            # Close the PDF response and return
            buffer.seek(0)
            response = HttpResponse(buffer.read(), content_type='application/pdf')
            response['Content-Disposition'] = 'attachment; filename="sum_example.pdf"'
            return response

    else:
        formset = ExcelSumFormSet()

    return render(request, 'excelapp/create_sum_pdf.html', {'formset': formset})


















def create_excel(request):
    if request.method == 'POST':
        formset = ExcelFormSet(request.POST)

        if formset.is_valid():
            # Create an Excel workbook
            workbook = Workbook()
            sheet = workbook.active

            # Write headers
            headers = ['Name', 'Age', 'City']
            for col_num, header in enumerate(headers):
                sheet.cell(row=1, column=col_num + 1, value=header)

            # Write data to the sheet and adjust column widths
            for row_num, form in enumerate(formset, start=2):
                data = [form.cleaned_data['name'], form.cleaned_data['age'], form.cleaned_data['city']]
                for col_num, value in enumerate(data):
                    cell = sheet.cell(row=row_num, column=col_num + 1, value=value)
                    column_letter = get_column_letter(col_num + 1)
                    column_width = len(str(value)) + 2
                    if sheet.column_dimensions[column_letter].width is None or sheet.column_dimensions[column_letter].width < column_width:
                        sheet.column_dimensions[column_letter].width = column_width  # Adjust column width

            # Set row heights based on content
            for row in sheet.iter_rows(min_row=2, max_row=len(formset) + 1):
                max_height = max(
                    cell.alignment.vertical if cell.alignment.vertical is not None else 0
                    for cell in row
                )
                sheet.row_dimensions[row[0].row].height = max_height

            # Create an HTTP response for the Excel file
            response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
            response['Content-Disposition'] = 'attachment; filename="example.xlsx"'
            workbook.save(response)

            return response
    else:
        formset = ExcelFormSet()

    return render(request, 'excelapp/create_excel.html', {'formset': formset})
